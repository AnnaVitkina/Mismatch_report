"""
Rate Costs Analysis Script

This script analyzes the rate.xlsx file from the "Rate card" tab:
1. Removes first 2 rows
2. Keeps first column (Lane #) + cost columns (from where costs start)
3. Renames cost columns based on row below: Flat -> "Price Flat", p/unit -> "Price per unit"
4. If MIN marker exists, adds " MIN" to the name
5. Creates CostColumn objects to store applies_if and rate_by conditions
"""

import pandas as pd
import openpyxl
import os
import re
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List


@dataclass
class CostColumn:
    """Represents a cost type with its conditions and pricing columns."""
    name: str                                    # e.g., "Pickup Fee"
    name_full: str                               # e.g., "Pickup Fee (Origin Collection Charges)"
    applies_if: Optional[str] = None             # e.g., "Applies if invoiced by Carrier"
    rate_by: Optional[str] = None                # e.g., "Rate by: Weight/chargeable kg"
    currency_col_idx: Optional[int] = None       # Column index for Currency
    price_columns: List[dict] = field(default_factory=list)  # [{"type": "Flat", "has_min": True, "col_idx": 15}, ...]
    
    def has_conditions(self) -> bool:
        """Check if this cost has any conditions."""
        return bool(self.applies_if or self.rate_by)
    
    def to_dict(self) -> dict:
        """Convert to dictionary for DataFrame export."""
        price_types = ", ".join([p.get("type", "") + (" MIN" if p.get("has_min") else "") 
                                  for p in self.price_columns])
        return {
            "Cost Name": self.name,
            "Full Name": self.name_full,
            "Applies If": self.applies_if or "",
            "Rate By": self.rate_by or "",
            "Currency Col Index": self.currency_col_idx,
            "Price Types": price_types,
            "Has Conditions": "Yes" if self.has_conditions() else "No"
        }


def get_agreement_number(file_path):
    """
    Extract Agreement number from "General info" tab of a rate card file.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        str: Agreement number or None if not found
    """
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    agreement_number = None
    try:
        workbook_info = openpyxl.load_workbook(full_path, data_only=True)
        if "General info" in workbook_info.sheetnames:
            general_info_sheet = workbook_info["General info"]
            # Find row with "Agreement number" in column A
            for row in general_info_sheet.iter_rows(min_col=1, max_col=2):
                cell_a = row[0]
                cell_b = row[1] if len(row) > 1 else None
                if cell_a.value and "Agreement number" in str(cell_a.value):
                    if cell_b and cell_b.value:
                        agreement_number = str(cell_b.value).strip()
                    break
        workbook_info.close()
    except Exception as e:
        print(f"   Warning: Could not extract Agreement number: {e}")
    
    return agreement_number


def load_rate_card_sheet(file_path):
    """Load the Rate card sheet from the Excel file using openpyxl."""
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    if "Rate card" not in workbook.sheetnames:
        raise ValueError(f"Sheet 'Rate card' not found in {file_path}")
    
    sheet = workbook["Rate card"]
    print(f"Loaded 'Rate card' sheet from {file_path}")
    print(f"   Total rows: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    
    return sheet, workbook


def get_row_values(sheet, row_number):
    """Get all values from a specific row (1-indexed)."""
    row_values = []
    for cell in sheet[row_number]:
        row_values.append(cell.value)
    return row_values


def clean_cost_name(value):
    """Clean cost name - just strip whitespace, keep parentheses."""
    if value is None:
        return None
    value_str = str(value).strip()
    return value_str if value_str else None


def find_first_cost_column(row_values):
    """Find the index of the first non-empty cost column."""
    for i, value in enumerate(row_values):
        if value is not None and str(value).strip() != '':
            return i
    return 0


def check_if_row_is_conditions_row(row_values):
    """
    Check if the row contains condition-related text (not cost names).
    
    A row is a conditions row (not cost names) if it contains:
    - "Applies if" or "Apply if"
    - "Condition" or "Conditional"
    - "Rate by"
    
    Returns:
        True if this row appears to be conditions (costs are in row above)
        False if this row appears to be cost names
    """
    condition_keywords = ['applies if', 'apply if', 'condition', 'rate by']
    
    for value in row_values:
        if value is not None:
            value_str = str(value).strip().lower()
            if not value_str:
                continue
            
            # Check if this cell starts with any condition keyword
            for keyword in condition_keywords:
                if value_str.startswith(keyword):
                    return True
    
    return False


def extract_cost_columns(sheet, first_cost_idx, cost_row, applies_if_row, rate_by_row, min_row, type_row):
    """
    Extract CostColumn objects for each cost type.
    
    Args:
        sheet: Worksheet object
        first_cost_idx: Index of first cost column
        cost_row: Row with cost names
        applies_if_row: Row with "Applies if..." conditions
        rate_by_row: Row with "Rate by:..." descriptions
        min_row: Row with MIN indicators
        type_row: Row with Currency/Flat/p/unit types
    
    Returns:
        List of CostColumn objects
    """
    cost_columns = []
    current_cost = None
    max_col = len(type_row)
    
    for col_idx in range(first_cost_idx, max_col):
        cost_name_full = cost_row[col_idx] if col_idx < len(cost_row) else None
        applies_if = applies_if_row[col_idx] if col_idx < len(applies_if_row) else None
        rate_by = rate_by_row[col_idx] if col_idx < len(rate_by_row) else None
        min_indicator = min_row[col_idx] if col_idx < len(min_row) else None
        col_type = type_row[col_idx] if col_idx < len(type_row) else None
        
        col_type_str = str(col_type).strip().lower() if col_type else ''
        min_str = str(min_indicator).strip().upper() if min_indicator else ''
        
        # Check if this is a new cost (Currency column with cost name)
        if col_type_str == 'currency' and cost_name_full and str(cost_name_full).strip():
            # Save previous cost if exists
            if current_cost is not None:
                cost_columns.append(current_cost)
            
            # Create new cost
            cost_name_clean = clean_cost_name(cost_name_full)
            current_cost = CostColumn(
                name=cost_name_clean,
                name_full=str(cost_name_full).strip(),
                applies_if=str(applies_if).strip() if applies_if and str(applies_if).strip() else None,
                rate_by=str(rate_by).strip() if rate_by and str(rate_by).strip() else None,
                currency_col_idx=col_idx,
                price_columns=[]
            )
        elif current_cost is not None:
            # Add price column to current cost
            if col_type_str == 'flat':
                current_cost.price_columns.append({
                    "type": "Flat",
                    "has_min": min_str == 'MIN',
                    "col_idx": col_idx
                })
            elif 'p/unit' in col_type_str or 'per unit' in col_type_str:
                current_cost.price_columns.append({
                    "type": "per unit",
                    "has_min": False,
                    "col_idx": col_idx
                })
    
    # Don't forget the last cost
    if current_cost is not None:
        cost_columns.append(current_cost)
    
    return cost_columns


def is_weight_range_value(value):
    """
    Check if a value looks like a weight range (e.g., "<= 200", ">= 100", "< 500", "> 1000").
    
    Returns:
        True if it's a weight range pattern
    """
    if value is None:
        return False
    val_str = str(value).strip()
    # Match patterns like: <= 200, >= 100, < 500, > 1000, =200
    return bool(re.match(r'^[<>=]+\s*\d+(\.\d+)?$', val_str))


def parse_weight_range(value):
    """
    Parse a weight range value like "<= 200" into (operator, number).
    
    Returns:
        tuple: (operator, number) or (None, None) if not a valid range
    """
    if value is None:
        return None, None
    val_str = str(value).strip()
    match = re.match(r'^([<>=]+)\s*(\d+(?:\.\d+)?)$', val_str)
    if match:
        return match.group(1), float(match.group(2))
    return None, None


def _build_weight_range_labels(range_values, weight_range_labels):
    """
    Build weight range labels for a group of consecutive columns.
    
    For ranges like: [(col1, "<=", 200), (col2, "<=", 500), (col3, "<=", 1000)]
    Creates labels: {col1: "<=200", col2: ">200 <=500", col3: ">500 <=1000"}
    
    Args:
        range_values: List of (col_idx, operator, number) tuples
        weight_range_labels: Dict to store {col_idx: label_string}
    """
    if not range_values:
        return
    
    # Sort by the number value to ensure proper ordering
    range_values_sorted = sorted(range_values, key=lambda x: x[2])
    
    prev_upper = None
    for i, (col_idx, op, num) in enumerate(range_values_sorted):
        # Format number (remove .0 for integers)
        num_str = str(int(num)) if num == int(num) else str(num)
        
        if prev_upper is None:
            # First range: just use the operator and number
            label = f"{op}{num_str}"
        else:
            # Subsequent ranges: add lower bound
            prev_str = str(int(prev_upper)) if prev_upper == int(prev_upper) else str(prev_upper)
            label = f">{prev_str} {op}{num_str}"
        
        weight_range_labels[col_idx] = label
        
        # Update prev_upper for next iteration
        if '<=' in op or '<' in op:
            prev_upper = num


def find_type_row(sheet, max_rows_to_check=15):
    """
    Find the row that contains column types (Currency, Flat, p/unit).
    This row indicates where the cost data structure is defined.
    
    Returns:
        tuple: (type_row_index, type_row_values, first_currency_col_idx)
    """
    for row_idx in range(3, min(max_rows_to_check + 1, sheet.max_row + 1)):
        row_values = get_row_values(sheet, row_idx)
        
        # Look for "Currency" in this row
        for col_idx, val in enumerate(row_values):
            if val is not None and str(val).strip().lower() == 'currency':
                print(f"   Found 'Currency' in Row {row_idx}, Column {col_idx}")
                return row_idx, row_values, col_idx
    
    return None, None, None


def create_filtered_dataframe(sheet):
    """
    Create a DataFrame keeping first column and cost columns with proper naming.
    
    DYNAMICALLY detects the structure by finding the row with "Currency", "Flat", "p/unit".
    Then works backwards to find cost names, applies_if, rate_by, and MIN rows.
    """
    # Step 1: Find the type row (contains Currency, Flat, p/unit)
    print("\n   Step 1: Finding type row (Currency, Flat, p/unit)...")
    type_row_idx, type_row, first_cost_idx = find_type_row(sheet)
    
    if type_row_idx is None:
        raise ValueError("Could not find type row (row with 'Currency' column)")
    
    print(f"   Type row found at Row {type_row_idx}")
    print(f"   First cost column index: {first_cost_idx}")
    
    # Step 2: Check the row above type row for MIN indicators AND/OR weight range indicators
    # NOTE: The same row can have MIN for some costs and weight ranges for others
    min_row_idx = type_row_idx - 1
    min_row = get_row_values(sheet, min_row_idx) if min_row_idx >= 1 else []
    
    # Check if min_row actually has MIN or MAX indicators
    has_min_indicators = any(
        val is not None and str(val).strip().upper() in ('MIN', 'MAX')
        for val in min_row
    )
    
    # Check if min_row has weight range indicators (e.g., "<= 200", "<= 500")
    has_weight_range_indicators = any(
        is_weight_range_value(val) for val in min_row
    )
    
    # Keep both - they can coexist for different cost types
    weight_range_row = min_row if has_weight_range_indicators else []
    
    if has_weight_range_indicators and has_min_indicators:
        print(f"   Row {min_row_idx} has BOTH MIN/MAX indicators and weight ranges")
        print(f"      Weight ranges: {[v for v in min_row if is_weight_range_value(v)][:5]}")
        print(f"      MIN/MAX: {[v for v in min_row if v and str(v).strip().upper() in ('MIN', 'MAX')][:5]}")
    elif has_weight_range_indicators:
        print(f"   Weight range row found at Row {min_row_idx}: {[v for v in min_row if is_weight_range_value(v)][:5]}")
    elif has_min_indicators:
        print(f"   MIN/MAX row found at Row {min_row_idx}")
    else:
        print(f"   No MIN/MAX/weight range indicators found in Row {min_row_idx}")
        min_row = []  # Empty - no MIN/MAX row
    
    # Step 3: Find cost names row and condition rows
    # Work backwards from type row to find cost names
    # Cost names row is the first row (going backwards) that has actual cost names
    # (not condition keywords like "Applies if", "Condition", "Rate by")
    
    cost_row_idx = None
    applies_if_row_idx = None
    rate_by_row_idx = None
    
    # Start checking from row above MIN (or type row if no MIN)
    start_check_row = min_row_idx - 1 if has_min_indicators else type_row_idx - 1
    
    for row_idx in range(start_check_row, 2, -1):  # Go backwards, stop at row 3
        row_values = get_row_values(sheet, row_idx)
        
        # Check if this row has condition keywords
        if check_if_row_is_conditions_row(row_values):
            # This is a conditions row
            first_val = None
            for val in row_values:
                if val is not None and str(val).strip():
                    first_val = str(val).strip().lower()
                    break
            
            if first_val:
                if first_val.startswith('rate by') or first_val.startswith('rate:'):
                    rate_by_row_idx = row_idx
                    print(f"   Rate By row found at Row {row_idx}")
                elif first_val.startswith('applies if') or first_val.startswith('apply if') or first_val.startswith('condition'):
                    applies_if_row_idx = row_idx
                    print(f"   Applies If row found at Row {row_idx}")
        else:
            # This could be the cost names row - check if it has non-empty values in cost columns
            has_cost_names = False
            for col_idx in range(first_cost_idx, len(row_values)):
                val = row_values[col_idx] if col_idx < len(row_values) else None
                if val is not None and str(val).strip():
                    has_cost_names = True
                    break
            
            if has_cost_names:
                cost_row_idx = row_idx
                print(f"   Cost names row found at Row {row_idx}")
                break
    
    if cost_row_idx is None:
        raise ValueError("Could not find cost names row")
    
    # Get the actual row values
    cost_row = get_row_values(sheet, cost_row_idx)
    applies_if_row = get_row_values(sheet, applies_if_row_idx) if applies_if_row_idx else []
    rate_by_row = get_row_values(sheet, rate_by_row_idx) if rate_by_row_idx else []
    
    # If we have weight ranges, we need to build range labels for consecutive columns
    # e.g., "<= 200", "<= 500", "<= 1000" becomes "<=200", ">200 <=500", ">500 <=1000"
    weight_range_labels = {}
    if weight_range_row:
        # Group consecutive weight ranges per cost type
        current_cost_start = None
        range_values = []
        
        for col_idx in range(first_cost_idx, len(weight_range_row)):
            val = weight_range_row[col_idx]
            cost_val = cost_row[col_idx] if col_idx < len(cost_row) else None
            
            # New cost type starts
            if cost_val and str(cost_val).strip():
                # Process previous group if exists
                if range_values:
                    _build_weight_range_labels(range_values, weight_range_labels)
                current_cost_start = col_idx
                range_values = []
            
            if is_weight_range_value(val):
                op, num = parse_weight_range(val)
                range_values.append((col_idx, op, num))
        
        # Don't forget the last group
        if range_values:
            _build_weight_range_labels(range_values, weight_range_labels)
        
        if weight_range_labels:
            print(f"   Built {len(weight_range_labels)} weight range column labels")
    
    # Debug output
    print(f"\n   Cost row (Row {cost_row_idx}, first 10 from col {first_cost_idx}): {cost_row[first_cost_idx:first_cost_idx+10]}")
    print(f"   Type row (Row {type_row_idx}, first 10 from col {first_cost_idx}): {type_row[first_cost_idx:first_cost_idx+10]}")
    
    # Step 4: Extract CostColumn objects
    cost_columns = extract_cost_columns(
        sheet, first_cost_idx, cost_row, applies_if_row, rate_by_row, min_row, type_row
    )
    print(f"\n   Extracted {len(cost_columns)} cost types")
    for i, cost in enumerate(cost_columns[:5]):
        print(f"      {i+1}. {cost.name}: applies_if={bool(cost.applies_if)}, rate_by={bool(cost.rate_by)}")
    if len(cost_columns) > 5:
        print(f"      ... and {len(cost_columns) - 5} more")
    
    # Step 5: Build DataFrame
    all_data = []
    for row in sheet.iter_rows(values_only=True):
        all_data.append(list(row))
    
    # Columns to keep: column 0 (Lane #) + columns from first_cost_idx onwards
    max_col = len(type_row)
    columns_to_keep = [0] + list(range(first_cost_idx, max_col))
    
    print(f"   Keeping {len(columns_to_keep)} columns (first col + cost cols)")
    
    # Build header names based on cost structure
    header = ['Lane #']  # First column is Lane #
    current_cost_name = None
    
    for col_idx in range(first_cost_idx, max_col):
        # Get values from different rows for this column
        cost_name = cost_row[col_idx] if col_idx < len(cost_row) else None
        min_indicator = min_row[col_idx] if col_idx < len(min_row) else None
        col_type = type_row[col_idx] if col_idx < len(type_row) else None
        weight_range_label = weight_range_labels.get(col_idx)
        
        # Clean cost name (remove parentheses)
        if cost_name is not None and str(cost_name).strip() != '':
            current_cost_name = clean_cost_name(cost_name)
        
        # Determine column header based on type
        col_type_str = str(col_type).strip().lower() if col_type else ''
        min_str = str(min_indicator).strip().upper() if min_indicator else ''
        
        # Check if this column has MIN or MAX indicator (not a weight range)
        is_min_column = min_str == 'MIN' and not weight_range_label
        is_max_column = min_str == 'MAX' and not weight_range_label
        
        if col_type_str == 'currency':
            # This is the currency column - use the cost name
            col_header = current_cost_name if current_cost_name else 'Currency'
        elif col_type_str == 'flat':
            # Price Flat column
            if is_min_column:
                col_header = 'Price Flat MIN'
            elif is_max_column:
                col_header = 'Price Flat MAX'
            elif weight_range_label:
                # Weight-tiered pricing: include weight range in column name
                col_header = f'Price Flat {weight_range_label}'
            else:
                col_header = 'Price Flat'
        elif 'p/unit' in col_type_str or 'per unit' in col_type_str:
            # Price per unit column
            if is_min_column:
                col_header = 'Price per unit MIN'
            elif is_max_column:
                col_header = 'Price per unit MAX'
            elif weight_range_label:
                col_header = f'Price per unit {weight_range_label}'
            else:
                col_header = 'Price per unit'
        else:
            # Unknown type - use the value from type row or cost name
            if col_type and str(col_type).strip():
                col_header = str(col_type).strip()
            elif current_cost_name:
                col_header = current_cost_name
            else:
                col_header = f'Column_{col_idx}'
        
        header.append(col_header)
    
    print(f"   Header (first 20): {header[:20]}")
    
    # Skip rows up to and including the type row (data starts after type row)
    data_start_idx = type_row_idx  # 0-indexed, so this is the first data row
    data_rows = all_data[data_start_idx:]
    
    # Filter columns for each row
    filtered_data = []
    for row in data_rows:
        filtered_row = [row[i] if i < len(row) else None for i in columns_to_keep]
        filtered_data.append(filtered_row)
    
    # Create DataFrame
    df = pd.DataFrame(filtered_data, columns=header)
    
    print(f"\n   Created DataFrame: {len(df)} rows x {len(df.columns)} columns")
    
    # Remove any header rows that might still be in the data
    # (rows where Lane # is not a number)
    rows_to_drop = []
    for idx, row in df.iterrows():
        lane_val = row['Lane #']
        if lane_val is None:
            rows_to_drop.append(idx)
            continue
        try:
            float(lane_val)
        except (ValueError, TypeError):
            rows_to_drop.append(idx)
    
    if rows_to_drop:
        print(f"\n   Removing {len(rows_to_drop)} non-data rows (Lane # not a number)")
        df = df.drop(index=rows_to_drop).reset_index(drop=True)
    
    print(f"   Final DataFrame: {len(df)} rows x {len(df.columns)} columns")
    
    return df, cost_columns


def save_to_excel(df, cost_columns, agreement_number=None, output_filename=None):
    """
    Save DataFrame and CostColumn info to Excel file in partly_df folder.
    
    Args:
        df: DataFrame with rate data
        cost_columns: List of CostColumn objects
        agreement_number: Optional agreement number for filename
        output_filename: Optional explicit output filename (overrides agreement_number)
    
    Returns:
        Path to the saved file
    """
    output_folder = Path(__file__).parent / "partly_df"
    output_folder.mkdir(exist_ok=True)
    
    # Determine output filename
    if output_filename:
        filename = output_filename
    elif agreement_number:
        # Use agreement number as filename
        safe_agreement = "".join(c for c in agreement_number if c.isalnum() or c in ('-', '_', ' ')).strip()
        filename = f"{safe_agreement}_costs.xlsx"
    else:
        filename = "rate_costs_filtered.xlsx"
    
    output_path = output_folder / filename
    
    # Create DataFrame from cost_columns
    cost_data = [cost.to_dict() for cost in cost_columns]
    df_costs = pd.DataFrame(cost_data)
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Rate Data', index=False)
            df_costs.to_excel(writer, sheet_name='Cost Conditions', index=False)
        print(f"\n   Saved to: {output_path}")
        print(f"      - Sheet 'Rate Data': {len(df)} rows")
        print(f"      - Sheet 'Cost Conditions': {len(df_costs)} cost types")
    except PermissionError:
        # File is open - try with alternative name
        alt_filename = filename.replace('.xlsx', '_new.xlsx')
        alt_path = output_folder / alt_filename
        with pd.ExcelWriter(alt_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Rate Data', index=False)
            df_costs.to_excel(writer, sheet_name='Cost Conditions', index=False)
        print(f"\n   [WARNING] Original file is open. Saved to: {alt_path}")
        print(f"      - Sheet 'Rate Data': {len(df)} rows")
        print(f"      - Sheet 'Cost Conditions': {len(df_costs)} cost types")
        output_path = alt_path
    
    return output_path


def process_single_rate_card(file_path):
    """
    Process a single rate card file and save output.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        tuple: (output_path, agreement_number) or (None, None) if error
    """
    print(f"\n   Processing: {file_path}")
    
    try:
        # Get agreement number
        agreement_number = get_agreement_number(file_path)
        if agreement_number:
            print(f"   Agreement number: {agreement_number}")
        else:
            print(f"   Agreement number: Not found")
        
        # Load the Rate card sheet
        sheet, workbook = load_rate_card_sheet(file_path)
        
        # Create filtered DataFrame and extract cost columns
        df, cost_columns = create_filtered_dataframe(sheet)
        
        # Save to Excel
        output_path = save_to_excel(df, cost_columns, agreement_number=agreement_number)
        
        # Close workbook
        workbook.close()
        
        return output_path, agreement_number
        
    except Exception as e:
        print(f"   [ERROR] Failed to process {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return None, None


def process_multiple_rate_cards(file_paths):
    """
    Process multiple rate card files and save each to a separate output file.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
                          (e.g., ["rate.xlsx", "rate_2.xlsx"])
    
    Returns:
        dict: Dictionary mapping agreement numbers to their output file paths
              {agreement_number: output_path, ...}
    """
    results = {}
    
    print(f"\n{'='*80}")
    print(f"RATE CARD COSTS ANALYSIS - Processing {len(file_paths)} rate card(s)")
    print(f"{'='*80}")
    
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] {'-'*60}")
        
        output_path, agreement_number = process_single_rate_card(file_path)
        
        if output_path:
            key = agreement_number if agreement_number else os.path.splitext(os.path.basename(file_path))[0]
            results[key] = str(output_path)
    
    print(f"\n{'='*80}")
    print(f"Processing complete! {len(results)}/{len(file_paths)} files processed successfully.")
    print(f"{'='*80}")
    
    if results:
        print("\nOutput files created:")
        for agreement, path in results.items():
            print(f"  - {agreement}: {path}")
    
    return results


def get_rate_card_files_from_input():
    """
    Get all Excel files from the input folder that could be rate cards.
    
    Returns:
        list: List of Excel file names in the input folder
    """
    input_folder = "input"
    if not os.path.exists(input_folder):
        print(f"Warning: Input folder '{input_folder}' does not exist.")
        return []
    
    excel_files = [f for f in os.listdir(input_folder) 
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    return excel_files


def main():
    """Main function to run the rate card costs analysis for multiple rate cards."""
    print("\n" + "="*80)
    print("RATE CARD COSTS ANALYSIS")
    print("="*80)
    
    # Process multiple rate cards (same as part4_rate_card_processing.py)
    rate_card_files = ["rate.xlsx", "rate_3.xlsx"]
    results = process_multiple_rate_cards(rate_card_files)
    
    return results


if __name__ == "__main__":
    results = main()




