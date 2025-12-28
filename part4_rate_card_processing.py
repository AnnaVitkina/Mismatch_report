import pandas as pd
import openpyxl
import os


def process_rate_card(file_path):
    """
    Process a Rate Card Excel file from the input folder.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder (e.g., "rate_card.xlsx")
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary, agreement number)
            - dataframe: Processed pandas DataFrame (filtered to black font columns)
            - list: List of column names in the processed dataframe
            - dict: Dictionary of conditions where keys are column names and values are condition text
            - str: Agreement number from "General Info" tab
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Extract Agreement number from "General Info" tab
    agreement_number = None
    try:
        workbook_info = openpyxl.load_workbook(full_path, data_only=True)
        if "General info" in workbook_info.sheetnames:
            general_info_sheet = workbook_info["General info"]
            # Find row with "Agreement number" in column A

            for row in general_info_sheet.iter_rows(min_col=1, max_col=2):
                print(row)
                cell_a = row[0]
                print(cell_a.value)
                cell_b = row[1] if len(row) > 1 else None
                print(cell_b.value)
                if cell_a.value and "Agreement number" in str(cell_a.value):
                    if cell_b and cell_b.value:
                        agreement_number = str(cell_b.value).strip()
                    break
        workbook_info.close()
    except Exception as e:
        print(f"Warning: Could not extract Agreement number from General Info tab: {e}")
    
    if agreement_number:
        print(f"   Agreement number: {agreement_number}")
    
    # Read the Excel file
    df_rate_card = pd.read_excel(full_path, sheet_name="Rate card", skiprows=2)
    
    # Find first column index (where data actually starts)
    first_column_index = None
    if df_rate_card is not None:
        for i, col in enumerate(df_rate_card.columns):
            if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                first_column_index = i
                break
    
    if first_column_index is not None:
        df_rate_card = df_rate_card.iloc[:, :first_column_index]
    
    # Drop rows where the first column is NaN
    if df_rate_card is not None:
        df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)
    
    # Set column names from first row
    new_columns = df_rate_card.iloc[0].tolist()
    df_rate_card.columns = new_columns
    df_rate_card = df_rate_card.iloc[1:]
    
    # Load the workbook to extract conditions and check font colors
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook["Rate card"]
    
    # Find the header row that contains "Currency"
    first_data_row_index = None
    currency_index = None
    
    for row_index in range(1, min(10, sheet.max_row + 1)):
        row = sheet[row_index]
        row_values = [cell.value for cell in row]
        if "Currency" in row_values:
            currency_index = row_values.index("Currency")
            first_data_row_index = row_index
            break
    
    black_font_values = []
    column_notes = {}  # Will store conditions/notes for each column
    
    if first_data_row_index is not None and currency_index is not None:
        # Access the data in this row
        first_data_row = sheet[first_data_row_index]
        first_data_values = [cell.value for cell in first_data_row]
        truncated_data_values = first_data_values[:currency_index]
        
        # Extract conditional rules/notes from multiple sources:
        # 1. Comments (notes) in the header row cells
        # 2. Cell values in the row ABOVE the header (row above column name)
        # 3. Cell values in row 2 (legacy fallback)
        header_row_index = first_data_row_index
        if header_row_index and header_row_index <= sheet.max_row:
            for i, col_name in enumerate(truncated_data_values, 1):
                if col_name:  # Only process non-empty column names
                    header_cell = sheet.cell(row=header_row_index, column=i)
                    
                    # Source 1: Check for comments (where conditional rules are stored)
                    if header_cell.comment:
                        comment_text = header_cell.comment.text
                        if comment_text and comment_text.strip():
                            column_notes[col_name] = comment_text.strip()
                    
                    # Source 2: Check the cell ABOVE the column name header
                    if col_name not in column_notes:
                        above_row_index = header_row_index - 1
                        if above_row_index >= 1:
                            above_cell = sheet.cell(row=above_row_index, column=i)
                            if above_cell.value and str(above_cell.value).strip():
                                column_notes[col_name] = str(above_cell.value).strip()
                    
                    # Source 3: Also check for cell value notes in row 2 (legacy fallback)
                    if col_name not in column_notes:
                        notes_row_index = 2
                        if notes_row_index <= sheet.max_row and notes_row_index != header_row_index - 1:
                            note_cell = sheet.cell(row=notes_row_index, column=i)
                            if note_cell.value and str(note_cell.value).strip():
                                column_notes[col_name] = str(note_cell.value).strip()
        
        # Check font color to identify black font columns (required columns)
        for i, value in enumerate(truncated_data_values):
            if i < len(first_data_row):
                cell = first_data_row[i]
                font_color = "black"
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color is not None:
                        # Convert to string and handle different formats
                        hex_str = str(hex_color).upper()
                        # Remove 'FF' prefix if present (ARGB format)
                        if hex_str.startswith('FF') and len(hex_str) == 8:
                            hex_str = hex_str[2:]
                        
                        # Check if it's black
                        if hex_str == '000000' or hex_str == '00000000':
                            font_color = "black"
                        else:
                            # Check if it's a shade of grey (R, G, and B are close)
                            try:
                                if len(hex_str) == 6:
                                    r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                                    # Check if it's a shade of grey (R, G, and B are close)
                                    if abs(r - g) < 10 and abs(g - b) < 10 and r > 0:  # Grey (not black, not white)
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"  # For colors that are not black or grey
                            except (ValueError, IndexError):
                                pass
                
                if font_color == "black":
                    black_font_values.append(value)
    
    # Filter the DataFrame to keep only the columns whose names are in black_font_values
    if df_rate_card is not None and black_font_values:
        # Only include columns that actually exist in the dataframe
        available_columns = [col for col in black_font_values if col in df_rate_card.columns]
        if available_columns:
            df_filtered_rate_card = df_rate_card[available_columns]
        else:
            df_filtered_rate_card = df_rate_card
    else:
        df_filtered_rate_card = df_rate_card
    
    # Get list of column names
    column_names = df_filtered_rate_card.columns.tolist()
    
    # Create conditions dictionary (only for columns that exist in the filtered dataframe)
    conditions = {}
    for col_name in column_names:
        if col_name in column_notes:
            conditions[col_name] = column_notes[col_name]
    
    return df_filtered_rate_card, column_names, conditions, agreement_number


def clean_condition_text(condition_text):
    """
    Clean up condition text for better readability.
    
    Transforms:
        "Conditional rules:
        1. 33321-6422: TOPOSTALCODE starts with 33321-6422,333216422"
    To:
        "1. 33321-6422: starts with 33321-6422,333216422"
    """
    import re
    
    if not condition_text:
        return condition_text
    
    # Remove "Conditional rules:" header (case insensitive)
    cleaned = re.sub(r'(?i)^conditional\s*rules\s*:\s*\n?', '', condition_text.strip())
    
    # Remove column name references like "TOPOSTALCODE ", "FROMPOSTALCODE ", etc.
    # Pattern: After the colon and value identifier, remove uppercase column names followed by space
    # Example: "33321-6422: TOPOSTALCODE starts with" -> "33321-6422: starts with"
    cleaned = re.sub(r':\s*[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r': \1', cleaned)
    
    # Also handle cases without numbered format
    cleaned = re.sub(r'^[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r'\1', cleaned, flags=re.MULTILINE)
    
    # Clean up extra whitespace and newlines
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    cleaned = '\n'.join(lines)
    
    return cleaned


def save_rate_card_output(file_path, output_path=None):
    """
    Process rate card and save output to Excel file with data and conditions.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
        output_path (str): Optional output path. If None, saves to "<agreement_number>.xlsx" in partly_df folder
    
    Returns:
        str: Path to the saved Excel file
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Process the rate card
    rate_card_dataframe, rate_card_column_names, rate_card_conditions, agreement_number = process_rate_card(file_path)
    
    # Set output path - save to partly_df folder (relative to script location)
    if output_path is None:
        # Get the directory where this script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Ensure partly_df folder exists in the script's directory
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        
        # Use agreement number as filename if available, otherwise use a default name
        if agreement_number:
            # Clean agreement number for use as filename (remove invalid characters)
            safe_agreement_number = "".join(c for c in agreement_number if c.isalnum() or c in ('-', '_', ' ')).strip()
            output_filename = f"{safe_agreement_number}.xlsx"
        else:
            # Fallback to original filename if no agreement number found
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{base_name}_processed.xlsx"
        
        output_path = os.path.join(partly_df_folder, output_filename)
    
    # Create conditions DataFrame with cleaned condition text
    conditions_data = []
    for col_name in rate_card_column_names:
        raw_condition = rate_card_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in rate_card_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Save to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Rate Card Data
        rate_card_dataframe.to_excel(writer, sheet_name='Rate Card Data', index=False)
        
        # Sheet 2: Conditions
        df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
        
        # Sheet 3: Summary
        summary_data = {
            'Metric': [
                'Agreement Number',
                'Total Rows',
                'Total Columns',
                'Columns with Conditions',
                'Columns without Conditions',
                'Source File'
            ],
            'Value': [
                agreement_number if agreement_number else 'Not found',
                len(rate_card_dataframe),
                len(rate_card_column_names),
                len(rate_card_conditions),
                len(rate_card_column_names) - len(rate_card_conditions),
                file_path
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply formatting
        workbook = writer.book
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        condition_yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        condition_no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format Rate Card Data sheet
        ws_data = workbook['Rate Card Data']
        for cell in ws_data[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_data.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        ws_data.freeze_panes = 'A2'
        
        # Format Conditions sheet
        ws_conditions = workbook['Conditions']
        for cell in ws_conditions[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color "Has Condition" column based on Yes/No
        for row in ws_conditions.iter_rows(min_row=2, max_row=ws_conditions.max_row):
            has_condition_cell = row[1]  # Column B (Has Condition)
            if has_condition_cell.value == 'Yes':
                has_condition_cell.fill = condition_yes_fill
            elif has_condition_cell.value == 'No':
                has_condition_cell.fill = condition_no_fill
            
            # Wrap text in Condition Rule column
            if len(row) > 2:
                row[2].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Set column widths for Conditions sheet
        ws_conditions.column_dimensions['A'].width = 30  # Column
        ws_conditions.column_dimensions['B'].width = 15  # Has Condition
        ws_conditions.column_dimensions['C'].width = 80  # Condition Rule
        
        ws_conditions.freeze_panes = 'A2'
        
        # Format Summary sheet
        ws_summary = workbook['Summary']
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 50
    
    print(f"\n✅ Rate Card output saved to: {output_path}")
    print(f"   - Agreement Number: {agreement_number if agreement_number else 'Not found'}")
    print(f"   - Sheet 'Rate Card Data': {len(rate_card_dataframe)} rows x {len(rate_card_column_names)} columns")
    print(f"   - Sheet 'Conditions': {len(rate_card_conditions)} columns with conditions")
    print(f"   - Sheet 'Summary': Overview statistics")
    
    return output_path


def process_multiple_rate_cards(file_paths):
    """
    Process multiple rate card files and save each to a separate output file.
    
    Args:
        file_paths (list): List of file paths relative to the "input/" folder
                          (e.g., ["rate_card_1.xlsx", "rate_card_2.xlsx"])
    
    Returns:
        dict: Dictionary mapping agreement numbers to their output file paths
              {agreement_number: output_path, ...}
    """
    results = {}
    
    print(f"\n{'='*60}")
    print(f"Processing {len(file_paths)} rate card(s)...")
    print(f"{'='*60}")
    
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] Processing: {file_path}")
        print("-" * 40)
        
        try:
            output_path = save_rate_card_output(file_path)
            
            # Get agreement number for the results dictionary
            _, _, _, agreement_number = process_rate_card(file_path)
            
            if agreement_number:
                results[agreement_number] = output_path
            else:
                # Use filename as key if no agreement number
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                results[base_name] = output_path
                
        except Exception as e:
            print(f"❌ Error processing {file_path}: {e}")
            continue
    
    print(f"\n{'='*60}")
    print(f"Processing complete! {len(results)}/{len(file_paths)} files processed successfully.")
    print(f"{'='*60}")
    
    if results:
        print("\nOutput files created:")
        for agreement, path in results.items():
            print(f"  - {agreement}: {path}")
    
    return results


def get_rate_card_files_from_input():
    """
    Get all Excel files from the input folder that could be rate cards.
    
    Returns:
        list: List of Excel file names in the input folder
    """
    input_folder = "input"
    if not os.path.exists(input_folder):
        print(f"Warning: Input folder '{input_folder}' does not exist.")
        return []
    
    excel_files = [f for f in os.listdir(input_folder) 
                   if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    
    return excel_files


if __name__ == "__main__":
    # Example 1: Process multiple specific rate cards
    rate_card_files = ["rate.xlsx", "rate_3.xlsx"]
    results = process_multiple_rate_cards(rate_card_files)
    
    # Example 2: Process a single rate card (backward compatible)
    # output_file = save_rate_card_output("rate.xlsx")
    
    #Example 3: Auto-discover and process all Excel files in input folder
    #all_rate_cards = get_rate_card_files_from_input()
    # if all_rate_cards:
    #     results = process_multiple_rate_cards(all_rate_cards)
    
    # Default behavior: Process single file for testing
    #print("Processing single rate card (rate.xlsx)...")
    #output_file = save_rate_card_output("rate.xlsx")
    
    # Also print details to console
    rate_card_dataframe, rate_card_column_names, rate_card_conditions, agreement_number = process_rate_card("rate.xlsx")
    print("\nAgreement Number:", agreement_number if agreement_number else "Not found")
    print("\nDataFrame shape:", rate_card_dataframe.shape)
    print("\nColumn names:")
    print(rate_card_column_names)
    print("\nConditions (cleaned):")
    for col, condition in rate_card_conditions.items():
        cleaned = clean_condition_text(condition)
        print(f"  {col}: {cleaned[:100]}..." if len(cleaned) > 100 else f"  {col}: {cleaned}")
